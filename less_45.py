import logging as lg
import re


lg.basicConfig(level=lg.ERROR, filename=r'C:\Users\USER\PycharmProjects\honework\שולחן העבודה\my_first_log',format='%(levelname)s:%(asctime)s:%(message)s')

lg.debug('this is debug')
lg.critical('this is critical')
lg.warning('this is warning')
lg.error("This is  error")



#---------------
try:
    with open(r'C:\Users\USER\PycharmProjects\honework\שולחן העבודה\my_first_log', 'r', encoding='utf-8') as file:
        content = file.read()

    match = re.search(r'[\d@#$%*]', content)

    if match:
       print(match.start())
    else:
       print(-1)

except FileNotFoundError:
    print("not found")


import openpyxl.chart as xlc
import openpyxl as xl

wb=xl.load_workbook(r'C:\Users\USER\Downloads\less5.xlsx')
sheet1=wb['גיליון1']

#print(sheet1.max_row)


with open(r"C:\Users\USER\PycharmProjects\honework\שולחן העבודה\my_first_log", "r", encoding="utf-8") as file:
    content = file.read()

t=0
for char in content:

    temp = sheet1.max_row + 1
    if char == '#':
        cell = sheet1[f"A{temp + 1}"]
        cell.value = t
    if char == '$':
        cell = sheet1[f"B{temp + 1}"]
        cell.value = t
    if char == '%':
        cell = sheet1[f"C{temp + 1}"]
        cell.value = t
    if char == '*':
        cell = sheet1[f"A{temp + 1}"]
        cell.value = t

    t+=1
    wb.save(r'C:\Users\USER\Downloads\less5.xlsx')


print("finish")